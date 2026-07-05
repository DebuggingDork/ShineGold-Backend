import csv
import io
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.enums import UserRole
from app.models.user import User
from app.repositories.user_repository import UserRepository
from app.schemas.user import BulkImportOut, BulkImportRowError, BulkImportedUser

# Column order in template and expected import file
TEMPLATE_HEADERS = ("employee_id", "name", "address", "mobile_number")
TEMPLATE_SAMPLE = ("EMP1001", "Ravi Kumar", "Hyderabad, Telangana", "9876543210")

INSTRUCTIONS = [
    ("Column", "Required", "Description"),
    ("employee_id", "Yes", "Unique login ID for the executive (e.g. EMP1001)"),
    ("name", "Yes", "Full name of the executive"),
    ("address", "Yes", "Office or base address (text only — used for records)"),
    ("mobile_number", "No", "Contact number"),
    ("", "", ""),
    ("Note", "", "Latitude/longitude are NOT entered in Excel."),
    ("", "", "After first login, the executive pins their location in the app."),
    ("", "", "That GPS point is used to sort farms by distance from home."),
]


class BulkImportError(Exception):
    pass


@dataclass
class ImportRow:
    row_number: int
    employee_id: str
    name: str
    address: str
    mobile_number: str | None = None


class BulkImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.user_repo = UserRepository(db)

    @staticmethod
    def build_template_bytes() -> bytes:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Executives"
        data_sheet.append(list(TEMPLATE_HEADERS))
        data_sheet.append(list(TEMPLATE_SAMPLE))

        guide = workbook.create_sheet("Instructions")
        for row in INSTRUCTIONS:
            guide.append(list(row))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def parse_upload(content: bytes, filename: str) -> list[ImportRow]:
        lowered = filename.lower()
        if lowered.endswith(".csv"):
            return BulkImportService._parse_csv(content)
        if lowered.endswith((".xlsx", ".xlsm")):
            return BulkImportService._parse_xlsx(content)
        raise BulkImportError("Unsupported file type. Upload a .xlsx or .csv file.")

    @staticmethod
    def _normalize_header(value: object | None) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_")

    @classmethod
    def _required_columns_present(cls, column_index: dict[str, int]) -> None:
        missing = [col for col in ("employee_id", "name", "address") if col not in column_index]
        if missing:
            raise BulkImportError(
                f"Missing required columns: {', '.join(missing)}. "
                f"Expected column order: employee_id, name, address, mobile_number (optional)"
            )

    @classmethod
    def _parse_xlsx(cls, content: bytes) -> list[ImportRow]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook["Executives"] if "Executives" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as e:
            raise BulkImportError("The uploaded file is empty.") from e

        headers = [cls._normalize_header(cell) for cell in header_row]
        column_index = {header: idx for idx, header in enumerate(headers) if header}
        cls._required_columns_present(column_index)

        parsed: list[ImportRow] = []
        for offset, row in enumerate(rows, start=2):
            if row is None or not any(row):
                continue

            employee_id = str(row[column_index["employee_id"]] or "").strip()
            name = str(row[column_index["name"]] or "").strip()
            address = str(row[column_index["address"]] or "").strip()
            mobile_number = None
            if "mobile_number" in column_index:
                raw_mobile = row[column_index["mobile_number"]]
                mobile_number = str(raw_mobile).strip() if raw_mobile not in (None, "") else None

            if not employee_id and not name and not address:
                continue
            parsed.append(
                ImportRow(
                    row_number=offset,
                    employee_id=employee_id,
                    name=name,
                    address=address,
                    mobile_number=mobile_number,
                )
            )
        workbook.close()
        return parsed

    @classmethod
    def _parse_csv(cls, content: bytes) -> list[ImportRow]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise BulkImportError("The uploaded CSV file is empty.")

        headers = {
            cls._normalize_header(name): name
            for name in reader.fieldnames
            if name
        }
        column_index = {key: 0 for key in headers}
        cls._required_columns_present(column_index)

        parsed: list[ImportRow] = []
        for offset, row in enumerate(reader, start=2):
            employee_id = (row.get(headers["employee_id"]) or "").strip()
            name = (row.get(headers["name"]) or "").strip()
            address = (row.get(headers["address"]) or "").strip()
            mobile_number = None
            if "mobile_number" in headers:
                raw_mobile = (row.get(headers["mobile_number"]) or "").strip()
                mobile_number = raw_mobile or None

            if not employee_id and not name and not address:
                continue
            parsed.append(
                ImportRow(
                    row_number=offset,
                    employee_id=employee_id,
                    name=name,
                    address=address,
                    mobile_number=mobile_number,
                )
            )
        return parsed

    async def import_executives(
        self,
        rows: list[ImportRow],
        default_password: str,
    ) -> BulkImportOut:
        if not rows:
            raise BulkImportError("No executive rows found in the uploaded file.")

        created_users: list[BulkImportedUser] = []
        errors: list[BulkImportRowError] = []
        skipped = 0
        seen_employee_ids: set[str] = set()

        for row in rows:
            if not row.employee_id or not row.name or not row.address:
                skipped += 1
                errors.append(
                    BulkImportRowError(
                        row=row.row_number,
                        employee_id=row.employee_id or None,
                        reason="employee_id, name, and address are required",
                    )
                )
                continue

            if row.employee_id in seen_employee_ids:
                skipped += 1
                errors.append(
                    BulkImportRowError(
                        row=row.row_number,
                        employee_id=row.employee_id,
                        reason="Duplicate employee_id in file",
                    )
                )
                continue
            seen_employee_ids.add(row.employee_id)

            existing = await self.user_repo.get_by_employee_id(row.employee_id)
            if existing is not None:
                skipped += 1
                errors.append(
                    BulkImportRowError(
                        row=row.row_number,
                        employee_id=row.employee_id,
                        reason="Employee ID already exists",
                    )
                )
                continue

            user = User(
                employee_id=row.employee_id,
                name=row.name,
                address=row.address,
                password_hash=hash_password(default_password),
                role=UserRole.EXECUTIVE,
                mobile_number=row.mobile_number,
            )
            created = await self.user_repo.create(user)
            created_users.append(
                BulkImportedUser(
                    id=created.id,
                    employee_id=created.employee_id,
                    name=created.name,
                    default_password=default_password,
                )
            )

        return BulkImportOut(
            created=len(created_users),
            skipped=skipped,
            errors=errors,
            users=created_users,
        )
